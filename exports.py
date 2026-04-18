"""
יצוא הביצועים לפורמטים פתוחים:

1. performance.xlsx — קובץ Excel עם 5 גיליונות (Summary, Signals, Forecasts,
   ModelAccuracy, Daily).
2. signals.csv — לוג כל האיתותים.
3. forecasts.csv — כל התחזיות והתצפיות המקבילות.
4. daily_performance.csv — סיכום ביצועים יומי.

הקבצים נוצרים מחדש בכל הרצה. מי שפותח אותם ישירות יקבל סנאפשוט. מי שמחבר
את Excel דרך Power Query לכתובת ה-CSV יקבל רענון אוטומטי.
"""
import csv
import datetime as dt
import json
import logging
import os
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import (
    CITIES, FORECASTS_LOG, OBSERVATIONS_JSON, OUTPUT_DIR,
    PRICES_LOG, SIGNALS_LOG, USER_TZ, WEATHER_MODELS,
)

log = logging.getLogger(__name__)

XLSX_PATH = f"{OUTPUT_DIR}/performance.xlsx"
SIGNALS_CSV = f"{OUTPUT_DIR}/signals.csv"
FORECASTS_CSV = f"{OUTPUT_DIR}/forecasts.csv"
DAILY_CSV = f"{OUTPUT_DIR}/daily_performance.csv"
PRICES_CSV = f"{OUTPUT_DIR}/prices.csv"


# ─────────────────────────────────────────────
# helpers
# ─────────────────────────────────────────────

def _load_jsonl(path: str) -> List[dict]:
    if not os.path.exists(path):
        return []
    rows = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return rows


def _load_json(path: str) -> dict:
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f) or {}
    except Exception:
        return {}


# ─────────────────────────────────────────────
# CSV exports — utf-8-sig לתמיכה בעברית ב-Excel
# ─────────────────────────────────────────────

def export_signals_csv() -> int:
    rows = _load_jsonl(SIGNALS_LOG)
    with open(SIGNALS_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "timestamp", "strategy", "city", "target_date", "action", "bucket",
            "bucket_type", "bucket_temp", "our_prob", "yes_price",
            "edge", "kelly", "ev", "stake_usd",
            "minutes_to_close", "timing",
            "status", "outcome_pnl", "observed_max", "settled_at",
        ])
        for r in rows:
            w.writerow([
                r.get("ts"), r.get("strategy") or "max_edge",
                r.get("city"), r.get("target_date"),
                r.get("action"), r.get("bucket_label"),
                r.get("bucket_type"), r.get("bucket_temp"),
                r.get("our_prob"), r.get("yes_price"),
                r.get("edge"), r.get("kelly"), r.get("ev"),
                r.get("stake_usd"),
                r.get("minutes_to_close"), r.get("timing"),
                r.get("status"), r.get("outcome_pnl"),
                r.get("observed_max"), r.get("settled_at"),
            ])
    return len(rows)


def export_forecasts_csv() -> int:
    forecasts = _load_jsonl(FORECASTS_LOG)
    obs = _load_json(OBSERVATIONS_JSON)

    # ממיינים: לכל (city, date) ניקח את התחזית האחרונה
    latest: Dict[tuple, dict] = {}
    for r in forecasts:
        key = (r.get("city") or "london", r.get("target_date"))
        if None in key:
            continue
        prev = latest.get(key)
        if prev is None or r.get("ts", "") > prev.get("ts", ""):
            latest[key] = r

    model_names = list(WEATHER_MODELS.keys())

    with open(FORECASTS_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "city", "target_date", "timestamp",
            *model_names, "consensus", "sigma",
            "observed_max", "consensus_error",
        ])
        for (city, date), r in sorted(latest.items()):
            city_obs = obs.get(city) or {}
            observed = city_obs.get(date)
            cons = r.get("consensus")
            err = (cons - observed) if (cons is not None and observed is not None) else None
            models = r.get("models") or {}
            w.writerow([
                city, date, r.get("ts"),
                *[models.get(m) for m in model_names],
                cons, r.get("sigma"),
                observed, err,
            ])
    return len(latest)


def export_prices_csv() -> int:
    rows = _load_jsonl(PRICES_LOG)
    with open(PRICES_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "timestamp", "city", "target_date",
            "bucket", "bucket_type", "bucket_temp",
            "yes_price", "yes_best_bid", "yes_best_ask",
            "volume", "our_prob", "minutes_to_close",
        ])
        for r in rows:
            w.writerow([
                r.get("ts"), r.get("city"), r.get("target_date"),
                r.get("bucket_label"), r.get("bucket_type"),
                r.get("bucket_temp"),
                r.get("yes_price"), r.get("yes_best_bid"),
                r.get("yes_best_ask"), r.get("volume"),
                r.get("our_prob"), r.get("minutes_to_close"),
            ])
    return len(rows)


def export_daily_performance_csv() -> int:
    """סיכום יומי לכל עיר: מספר איתותים, רווח, אחוז זכייה."""
    signals = _load_jsonl(SIGNALS_LOG)
    # קיבוץ לפי (date, city)
    buckets: Dict[tuple, dict] = {}
    for s in signals:
        date = s.get("target_date")
        city = s.get("city")
        if not date or not city:
            continue
        key = (date, city)
        b = buckets.setdefault(key, {
            "signals": 0, "won": 0, "lost": 0, "pending": 0,
            "pnl": 0.0, "stake": 0.0,
        })
        b["signals"] += 1
        b["stake"]   += float(s.get("stake_usd") or 0)
        st = s.get("status") or "pending"
        if st == "won":
            b["won"] += 1
            b["pnl"] += float(s.get("outcome_pnl") or 0)
        elif st == "lost":
            b["lost"] += 1
            b["pnl"] += float(s.get("outcome_pnl") or 0)
        else:
            b["pending"] += 1

    with open(DAILY_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["target_date", "city", "signals", "won", "lost",
                    "pending", "stake_usd", "pnl_usd", "win_rate"])
        for (date, city) in sorted(buckets.keys()):
            b = buckets[(date, city)]
            settled = b["won"] + b["lost"]
            win_rate = (b["won"] / settled) if settled > 0 else None
            w.writerow([date, city, b["signals"], b["won"], b["lost"],
                        b["pending"], round(b["stake"], 2),
                        round(b["pnl"], 2),
                        round(win_rate, 4) if win_rate is not None else None])
    return len(buckets)


# ─────────────────────────────────────────────
# XLSX — קובץ Excel שלם עם 5 גיליונות
# ─────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F2932")
_HEADER_FONT = Font(color="B5EBBF", bold=True)
_WON_FILL    = PatternFill("solid", fgColor="1A3A22")
_LOST_FILL   = PatternFill("solid", fgColor="3A1A1A")
_PEND_FILL   = PatternFill("solid", fgColor="2A2A1A")


def _style_header_row(ws, row: int, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws, min_w: int = 10, max_w: int = 30):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = min_w
        for cell in col:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, min(max_w, len(str(v))))
        ws.column_dimensions[col_letter].width = max_len + 2


def _sheet_summary(ws, performance: dict, accuracy: dict, generated_at: str):
    ws["A1"] = "סיכום ביצועים"
    ws["A1"].font = Font(size=16, bold=True, color="B5EBBF")
    ws.merge_cells("A1:D1")
    ws["A2"] = "רוענן לאחרונה:"
    ws["B2"] = generated_at

    row = 4
    ws.cell(row=row, column=1, value="מדד").font = Font(bold=True)
    ws.cell(row=row, column=2, value="ערך").font = Font(bold=True)
    _style_header_row(ws, row, 2)
    row += 1

    kpis = [
        ("איתותים סה״כ",        performance.get("total", 0)),
        ("פתוחים",               performance.get("pending", 0)),
        ("סגורים",               performance.get("settled", 0)),
        ("זכיות",                performance.get("won", 0)),
        ("הפסדים",               performance.get("lost", 0)),
        ("אחוז זכייה",           performance.get("win_rate")),
        ("רווח ממומש (דולר)",    performance.get("realized_pnl", 0)),
        ("רווח צפוי (EV) (דולר)", performance.get("expected_pnl", 0)),
        ("סטייק פתוח (דולר)",    performance.get("pending_stake", 0)),
        ("ROI על סטייק סגור",    performance.get("roi_on_settled_stake")),
    ]
    for label, value in kpis:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # פירוט לפי אסטרטגיה
    row += 1
    ws.cell(row=row, column=1, value="ביצועים לפי אסטרטגיה").font = Font(size=13, bold=True, color="B5EBBF")
    row += 1
    headers_s = ["אסטרטגיה", "סה״כ", "זכיות", "הפסדים", "פתוחים", "אחוז זכייה", "P&L ($)", "ROI"]
    for i, h in enumerate(headers_s, start=1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(headers_s))
    row += 1
    names_he = {"max_edge": "יתרון מקסימלי", "most_likely": "הסביר ביותר"}
    for s_name, s in (performance.get("by_strategy") or {}).items():
        if (s.get("total") or 0) == 0:
            continue
        ws.cell(row=row, column=1, value=names_he.get(s_name, s_name))
        ws.cell(row=row, column=2, value=s.get("total", 0))
        ws.cell(row=row, column=3, value=s.get("won", 0))
        ws.cell(row=row, column=4, value=s.get("lost", 0))
        ws.cell(row=row, column=5, value=s.get("pending", 0))
        ws.cell(row=row, column=6, value=s.get("win_rate"))
        ws.cell(row=row, column=7, value=s.get("realized_pnl"))
        ws.cell(row=row, column=8, value=s.get("roi_on_settled_stake"))
        row += 1

    # פירוט לפי עיר
    row += 1
    ws.cell(row=row, column=1, value="ביצועים לפי עיר (כל האסטרטגיות)").font = Font(size=13, bold=True, color="B5EBBF")
    row += 1
    headers = ["עיר", "סה״כ", "זכיות", "הפסדים", "פתוחים", "רווח (דולר)"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1
    for city_key, slot in (performance.get("per_city") or {}).items():
        ws.cell(row=row, column=1, value=city_key)
        ws.cell(row=row, column=2, value=slot.get("total", 0))
        ws.cell(row=row, column=3, value=slot.get("won", 0))
        ws.cell(row=row, column=4, value=slot.get("lost", 0))
        ws.cell(row=row, column=5, value=slot.get("pending", 0))
        ws.cell(row=row, column=6, value=round(slot.get("pnl") or 0, 2))
        row += 1

    _autosize(ws)


def _sheet_signals(ws):
    rows = _load_jsonl(SIGNALS_LOG)
    headers = [
        "תאריך-שעה הרישום", "אסטרטגיה", "עיר", "תאריך יעד", "פעולה", "bucket",
        "הסתברות שלנו", "מחיר שוק YES", "יתרון", "Kelly", "EV",
        "סטייק ($)", "דק׳ לסגירה", "קטגוריית תזמון",
        "סטטוס", "רווח/הפסד ($)", "טמפ' שנמדדה",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1,  value=r.get("ts"))
        ws.cell(row=i, column=2,  value=r.get("strategy") or "max_edge")
        ws.cell(row=i, column=3,  value=r.get("city"))
        ws.cell(row=i, column=4,  value=r.get("target_date"))
        ws.cell(row=i, column=5,  value=r.get("action"))
        ws.cell(row=i, column=6,  value=r.get("bucket_label"))
        ws.cell(row=i, column=7,  value=r.get("our_prob"))
        ws.cell(row=i, column=8,  value=r.get("yes_price"))
        ws.cell(row=i, column=9,  value=r.get("edge"))
        ws.cell(row=i, column=10, value=r.get("kelly"))
        ws.cell(row=i, column=11, value=r.get("ev"))
        ws.cell(row=i, column=12, value=r.get("stake_usd"))
        ws.cell(row=i, column=13, value=r.get("minutes_to_close"))
        ws.cell(row=i, column=14, value=r.get("timing"))
        ws.cell(row=i, column=15, value=r.get("status"))
        ws.cell(row=i, column=16, value=r.get("outcome_pnl"))
        ws.cell(row=i, column=17, value=r.get("observed_max"))

        st = r.get("status") or "pending"
        fill = _WON_FILL if st == "won" else (_LOST_FILL if st == "lost" else _PEND_FILL)
        for c in range(1, len(headers) + 1):
            ws.cell(row=i, column=c).fill = fill

    _autosize(ws)


def _sheet_forecasts(ws):
    forecasts = _load_jsonl(FORECASTS_LOG)
    obs = _load_json(OBSERVATIONS_JSON)
    latest: Dict[tuple, dict] = {}
    for r in forecasts:
        key = (r.get("city") or "london", r.get("target_date"))
        if None in key:
            continue
        prev = latest.get(key)
        if prev is None or r.get("ts", "") > prev.get("ts", ""):
            latest[key] = r

    model_names = list(WEATHER_MODELS.keys())
    headers = ["עיר", "תאריך יעד", "רוענן לאחרונה",
               *model_names, "קונצנזוס", "σ מודלים",
               "נמדד בפועל", "שגיאת קונצנזוס"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    row = 2
    for (city, date), r in sorted(latest.items()):
        city_obs = obs.get(city) or {}
        observed = city_obs.get(date)
        cons = r.get("consensus")
        err = (cons - observed) if (cons is not None and observed is not None) else None
        models = r.get("models") or {}
        ws.cell(row=row, column=1, value=city)
        ws.cell(row=row, column=2, value=date)
        ws.cell(row=row, column=3, value=r.get("ts"))
        for i, m in enumerate(model_names, start=4):
            ws.cell(row=row, column=i, value=models.get(m))
        ws.cell(row=row, column=4 + len(model_names), value=cons)
        ws.cell(row=row, column=5 + len(model_names), value=r.get("sigma"))
        ws.cell(row=row, column=6 + len(model_names), value=observed)
        ws.cell(row=row, column=7 + len(model_names), value=err)
        row += 1

    _autosize(ws)


def _sheet_accuracy(ws, accuracy: dict):
    if not accuracy:
        ws.cell(row=1, column=1, value="לא נצברה עדיין היסטוריה.")
        return

    row = 1
    ws.cell(row=row, column=1, value="איכות מודלים — מצטבר גלובלי").font = Font(size=13, bold=True, color="B5EBBF")
    row += 1

    headers = ["מודל", "MAE (°C)", "הטיה (°C)", "פגיעה ב-±1°C",
               "פגיעה ב-bucket", "דירוג ממוצע", "ימים"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(headers))
    row += 1

    glob = accuracy.get("global") or {}
    for name, s in (glob.get("models") or {}).items():
        if (s.get("n") or 0) == 0:
            continue
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=round(s["mae"], 3) if s.get("mae") is not None else None)
        ws.cell(row=row, column=3, value=round(s["bias"], 3) if s.get("bias") is not None else None)
        ws.cell(row=row, column=4, value=round(s["hit_1c"], 3) if s.get("hit_1c") is not None else None)
        ws.cell(row=row, column=5, value=round(s["bucket_hit"], 3) if s.get("bucket_hit") is not None else None)
        ws.cell(row=row, column=6, value=round(s["rank_avg"], 2) if s.get("rank_avg") is not None else None)
        ws.cell(row=row, column=7, value=s.get("n"))
        row += 1

    cons = glob.get("consensus") or {}
    if cons.get("n"):
        ws.cell(row=row, column=1, value="קונצנזוס").font = Font(bold=True)
        ws.cell(row=row, column=2, value=round(cons["mae"], 3) if cons.get("mae") is not None else None)
        ws.cell(row=row, column=3, value=round(cons["bias"], 3) if cons.get("bias") is not None else None)
        ws.cell(row=row, column=4, value=round(cons["hit_1c"], 3) if cons.get("hit_1c") is not None else None)
        ws.cell(row=row, column=5, value=round(cons["bucket_hit"], 3) if cons.get("bucket_hit") is not None else None)
        ws.cell(row=row, column=7, value=cons.get("n"))
        row += 1

    # פירוט לפי עיר
    row += 2
    ws.cell(row=row, column=1, value="איכות מודלים — לפי עיר").font = Font(size=13, bold=True, color="B5EBBF")
    row += 1
    headers2 = ["עיר", "מודל", "MAE", "הטיה", "פגיעה ב-±1°C",
                "פגיעה ב-bucket", "דירוג ממוצע", "ימים"]
    for i, h in enumerate(headers2, start=1):
        ws.cell(row=row, column=i, value=h)
    _style_header_row(ws, row, len(headers2))
    row += 1
    for city_key, city_scores in (accuracy.get("per_city") or {}).items():
        for name, s in (city_scores.get("models") or {}).items():
            if (s.get("n") or 0) == 0:
                continue
            ws.cell(row=row, column=1, value=city_key)
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=round(s["mae"], 3))
            ws.cell(row=row, column=4, value=round(s["bias"], 3))
            ws.cell(row=row, column=5, value=round(s["hit_1c"], 3))
            ws.cell(row=row, column=6, value=round(s["bucket_hit"], 3) if s.get("bucket_hit") is not None else None)
            ws.cell(row=row, column=7, value=round(s["rank_avg"], 2) if s.get("rank_avg") is not None else None)
            ws.cell(row=row, column=8, value=s.get("n"))
            row += 1

    _autosize(ws)


def _sheet_daily(ws):
    signals = _load_jsonl(SIGNALS_LOG)
    buckets: Dict[tuple, dict] = {}
    for s in signals:
        date = s.get("target_date")
        city = s.get("city")
        if not date or not city:
            continue
        key = (date, city)
        b = buckets.setdefault(key, {
            "signals": 0, "won": 0, "lost": 0, "pending": 0,
            "pnl": 0.0, "stake": 0.0,
        })
        b["signals"] += 1
        b["stake"]   += float(s.get("stake_usd") or 0)
        st = s.get("status") or "pending"
        if st == "won":
            b["won"] += 1
            b["pnl"] += float(s.get("outcome_pnl") or 0)
        elif st == "lost":
            b["lost"] += 1
            b["pnl"] += float(s.get("outcome_pnl") or 0)
        else:
            b["pending"] += 1

    headers = ["תאריך יעד", "עיר", "איתותים", "זכיות", "הפסדים",
               "פתוחים", "סטייק ($)", "רווח/הפסד ($)", "אחוז זכייה"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    _style_header_row(ws, 1, len(headers))

    row = 2
    for (date, city) in sorted(buckets.keys()):
        b = buckets[(date, city)]
        settled = b["won"] + b["lost"]
        win_rate = (b["won"] / settled) if settled > 0 else None
        ws.cell(row=row, column=1, value=date)
        ws.cell(row=row, column=2, value=city)
        ws.cell(row=row, column=3, value=b["signals"])
        ws.cell(row=row, column=4, value=b["won"])
        ws.cell(row=row, column=5, value=b["lost"])
        ws.cell(row=row, column=6, value=b["pending"])
        ws.cell(row=row, column=7, value=round(b["stake"], 2))
        ws.cell(row=row, column=8, value=round(b["pnl"], 2))
        ws.cell(row=row, column=9, value=round(win_rate, 4) if win_rate is not None else None)
        row += 1

    _autosize(ws)


def export_xlsx(performance: dict, accuracy: dict,
                generated_at: str) -> None:
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "סיכום"
    ws_summary.sheet_view.rightToLeft = True
    _sheet_summary(ws_summary, performance or {}, accuracy or {}, generated_at)

    ws_signals = wb.create_sheet("איתותים")
    ws_signals.sheet_view.rightToLeft = True
    _sheet_signals(ws_signals)

    ws_forecasts = wb.create_sheet("תחזיות ותצפיות")
    ws_forecasts.sheet_view.rightToLeft = True
    _sheet_forecasts(ws_forecasts)

    ws_accuracy = wb.create_sheet("איכות מודלים")
    ws_accuracy.sheet_view.rightToLeft = True
    _sheet_accuracy(ws_accuracy, accuracy or {})

    ws_daily = wb.create_sheet("יומי")
    ws_daily.sheet_view.rightToLeft = True
    _sheet_daily(ws_daily)

    wb.save(XLSX_PATH)


def export_all(performance: dict, accuracy: dict,
               generated_at: str) -> dict:
    n_signals   = export_signals_csv()
    n_forecasts = export_forecasts_csv()
    n_daily     = export_daily_performance_csv()
    n_prices    = export_prices_csv()
    try:
        export_xlsx(performance, accuracy, generated_at)
    except Exception as e:
        log.warning("יצוא XLSX נכשל: %s", e)
    log.info("יצוא: %d איתותים, %d תחזיות, %d ימים, %d מחירים → %s",
             n_signals, n_forecasts, n_daily, n_prices, XLSX_PATH)
    return {
        "signals_csv":   SIGNALS_CSV,
        "forecasts_csv": FORECASTS_CSV,
        "daily_csv":     DAILY_CSV,
        "prices_csv":    PRICES_CSV,
        "xlsx":          XLSX_PATH,
    }
